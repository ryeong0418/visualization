from src import common_module as cm
from src.common.constants import SystemConstants, DbTypeConstants
from src.common.utils import DateUtils, SystemUtils, SqlUtils, ExcelUtils
from src.analysis_extend_target import OracleTarget

import os
import re
import pandas as pd
from openpyxl import load_workbook
from resources.config_manager import Config
from openpyxl.chart import Reference
import warnings


class MetricPerformanceReport(cm.CommonModule):
    """
    AE_DB_SYSMETRIC 테이블에 있는 데이터 가져와서
    EXCEL에 데이터 및 그래프 INSERT
    """

    def __init__(self, logger):
        super().__init__(logger=logger)
        self.ot: OracleTarget = None

    def main_process(self):
        """
        metric performance report 모듈을 실행하기 위한 main processd
        """

        self.logger.info("metric performance report")
        warnings.filterwarnings(action='ignore')
        self._insert_extend_target_data()
        self._insert_datatable_or_chartgraph()

    def _insert_extend_target_data(self):
        """
        DB 확장 타겟(실제 분석 대상 DB) 데이터 저장 함수
        """
        extend_target_repo_list = self.config["maxgauge_repo"].get("extend_target_repo", [])

        for extend_target_repo in extend_target_repo_list:
            extend_target_repo["analysis_target_type"] = self.config["maxgauge_repo"]["analysis_target_type"]  # oracle

            if str(self.config["maxgauge_repo"]["analysis_target_type"]).lower() == DbTypeConstants.ORACLE:
                if self.ot is None:
                    self.ot = OracleTarget(self.logger, self.config)

                self.ot.set_extend_target_config(extend_target_repo)
                self.ot.init_process()

    def _convert_sql_to_df(self, sql_path, filename):
        """
        sql query문을 dataframe 형태로 변환
        """
        metric_name_list = self._extract_metric_name_list()
        unpack_metric_name_list = str(metric_name_list)[1:-1]

        s_date, e_date = DateUtils.get_each_date_by_interval2(
            self.config["args"]["s_date"], self.config["args"]["interval"], arg_fmt="%Y-%m-%d"
        )
        date_dict = {"StartDate": s_date, "EndDate": e_date, "Metric_Name": unpack_metric_name_list}

        query = SystemUtils.get_file_content_in_path(sql_path, filename + ".txt")
        date_query = SqlUtils.sql_replace_to_dict(query, date_dict)

        for df in self.ot.get_data_by_query(date_query):
            df.columns = [i.upper() for i in df.columns]
            return df

    def _read_excel(self, excel_file_path):
        wb = load_workbook(excel_file_path)
        return wb

    def _extract_metric_name_list(self):
        metric_name_list = [
            "Host CPU Utilization (%)",
            "Average Active Sessions",
            "Executions Per Sec",
            "User Transaction Per Sec",
            "Logical Reads Per Sec",
            "Physical Reads Per Sec",
            "Hard Parse Count Per Sec",
        ]

        config_report = Config("report").get_config()
        metric_name_list.extend(config_report["sys_metric"])

        return metric_name_list

    def _make_union_df(self, df, metric_name):
        metric_dic = {}
        metric_value_list = []

        metric_df = df[df["METRIC_NAME"] == metric_name]

        metric_df["DATE_TIME"] = metric_df["DATE_TIME"].astype(str) + ":00"
        metric_df["DATE_TIME"] = metric_df["DATE_TIME"].apply(lambda x: "\n".join(x.split()))

        for idx, instance_num in enumerate(metric_df["INSTANCE_NUMBER"].unique()):
            extract_df = metric_df[metric_df["INSTANCE_NUMBER"] == instance_num]
            variable_name = f"extract_df_{instance_num}"
            metric_dic[variable_name] = extract_df.rename(
                columns={
                    "AG": f"AVG_{instance_num}",
                    "INSTANCE_NUMBER": f"INSTANCE_NUMBER_{instance_num}",
                    "MX": f"MAX_{instance_num}",
                }
            )

            metric_value_list = list(metric_dic.values())

        result_df = (
            pd.DataFrame()
            .join([d.set_index(["DATE_TIME", "METRIC_NAME"]) for d in metric_value_list], how="outer")
            .reset_index()
        )

        return result_df

    def _insert_datatable_or_chartgraph(self):
        """
        column txt 파일 읽어서 column명 추출
        {number} txt 파일명 추출
        """
        sql_path = f"{self.config['home']}/" + SystemConstants.CHART_SQL
        excel_path = f"{self.config['home']}/" + SystemConstants.CHART_EXCEL
        os.makedirs(excel_path, exist_ok=True)
        txt_file_list = SystemUtils.get_filenames_from_path(sql_path)

        for file in txt_file_list:
            filename = file.split(".")[0]
            excel_file_path = f"{excel_path}/{filename}.xlsx"

            if re.search(r"-(\d+)", filename):
                df = self._convert_sql_to_df(sql_path, filename)
                ExcelUtils.excel_export(excel_file_path, "Sheet1", df)

            elif re.search(r"CHART", filename):
                self.check_excel_format(excel_file_path, sql_path, filename)

    def check_excel_format(self, excel_file_path, sql_path, filename):
        """
        excel 파일이 있으면 해당 파일에 데이터 overwrite하고
        excel 파일이 없으면 새로 생성하여 데이터 insert 한다.
        """
        if os.path.isfile(excel_file_path):
            #self.logger.info(f"{filename}: filename OVERWRITE")
            self._check_sheet_name_list(excel_file_path, sql_path, filename)
            self._apply_excel_style(excel_file_path, sql_path, filename)
            self._insert_linechart_from_data(excel_file_path)

        else:
            #self.logger.info(f"{filename}: filename INSERT")
            self._make_excel_sheet_data(excel_file_path, sql_path, filename)
            self._apply_excel_style(excel_file_path, sql_path, filename)
            self._insert_linechart_from_data(excel_file_path)

    def _make_excel_sheet_data(self, excel_file_path, sql_path, filename):
        df = self._convert_sql_to_df(sql_path, filename)
        metric_name_list = df["METRIC_NAME"].unique()
        ExcelUtils.create_excel_and_sheet(excel_file_path, metric_name_list)
        self._insert_df_into_excel(excel_file_path, df, metric_name_list)

    def _insert_df_into_excel(self, excel_file_path, df, metric_name_list):
        for metric_name in metric_name_list:
            result_df = self._make_union_df(df, metric_name)
            ExcelUtils.append_df_into_excel(excel_file_path, metric_name, result_df, 2, 29, "overlay")

    def _check_sheet_name_list(self, excel_file_path, sql_path, filename):
        """
        sheet_name check
        """
        wb = self._read_excel(excel_file_path)
        df = self._convert_sql_to_df(sql_path, filename)
        metric_name_list = self._extract_metric_name_list()

        not_exist_sheet_list = []
        exist_sheet_list = []

        for sheet_name in metric_name_list:
            if sheet_name in wb.get_sheet_names():
                exist_sheet_list.append(sheet_name)
            else:
                not_exist_sheet_list.append(sheet_name)

        wb.close()

        if len(not_exist_sheet_list) != 0:
            #self.logger.info(f"{filename}: filename, {sheet_name} : not_exist_sheet")
            self._insert_df_into_excel(excel_file_path, df, not_exist_sheet_list)

        #self.logger.info(f"{filename} overwrite START!")
        self._overwrite_excel_sheet(exist_sheet_list, excel_file_path, df)

    def _overwrite_excel_sheet(self, sheet_name_list, excel_file_path, df):
        for sheet_name in sheet_name_list:
            wb = self._read_excel(excel_file_path)
            ws = wb[sheet_name]
            wb.close()

            result_df = self._make_union_df(df, sheet_name)
            col = [cell for cell in ws[ws.min_row] if cell.value == "DATE_TIME"]

            for date_time_cell in col:
                ExcelUtils.append_df_into_excel(
                    excel_file_path, sheet_name, result_df, date_time_cell.column - 1, ws.min_row - 1, "replace"
                )

    def _apply_excel_style(self, excel_file_path, sql_path, filename):
        """
        excel에 dataframe 기입시 스타일 지정
        table border_style, column width 지정
        """
        wb = self._read_excel(excel_file_path)
        df = self._convert_sql_to_df(sql_path, filename)
        unique_metric_name = df["METRIC_NAME"].unique()

        for metric_name in unique_metric_name:
            ws = wb[metric_name]
            SystemUtils.apply_thin_border(ws, "thin")
            SystemUtils.apply_column_width(ws, 20)

        wb.save(excel_file_path)
        wb.close()

    def _insert_linechart_from_data(self, excel_file_path):
        wb = self._read_excel(excel_file_path)
        metric_name_list = self._extract_metric_name_list()

        for metric_name in metric_name_list:
            ws = wb[metric_name]

            avg_col = [cell for cell in ws[ws.min_row] if isinstance(cell.value, str) and re.search(r"AVG", cell.value)]
            max_col = [cell for cell in ws[ws.min_row] if isinstance(cell.value, str) and re.search(r"MAX", cell.value)]

            category = Reference(
                ws, min_col=ws.min_column + 2, max_col=ws.min_column + 2, min_row=ws.min_row + 1, max_row=ws.max_row
            )

            AVG_METRIC_NAME = f"{metric_name}-AVG"
            line_chart_avg = ExcelUtils.set_linechart_object(AVG_METRIC_NAME)
            ExcelUtils.set_data_and_category(ws, category, avg_col, line_chart_avg)
            ExcelUtils.set_series_marker_style(line_chart_avg.series)
            ws.add_chart(line_chart_avg, "C1")

            MAX_METRIC_NAME = f"{metric_name}-MAX"
            line_chart_max = ExcelUtils.set_linechart_object(MAX_METRIC_NAME)
            ExcelUtils.set_data_and_category(ws, category, max_col, line_chart_max)
            ExcelUtils.set_series_marker_style(line_chart_max.series)
            ws.add_chart(line_chart_max, "C15")

        wb.save(excel_file_path)
        wb.close()
